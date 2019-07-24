'''
Created on Feb 15, 2019

@author: binakarir
'''
from openpyxl import load_workbook, Workbook
import pathlib
import os



class PyExcel():

    def __init__(self, *args):
        self.data = args   
        # print (self.data) 
        self.pathexcel = pathlib.Path.cwd()/"models/aaa.xlsx"
        # print (str(self.pathexcel))

        self.wb = load_workbook(filename = self.pathexcel, read_only = False)
        # ws =  wb.active
        self.ws = self.wb.worksheets[0]
        # ws['A1'] = 'E'
        # ws["C3"] = 'C' 
        self.__insertdata()
        self.wb.save(self.pathexcel)

    def __insertdata(self):
        col = 1
        j = 0
        for j in range(0, 9):
            row = 1
            for i in self.data[j]:
                self.ws.cell(row = row, column = col).value = i
                row += 1
            col += 1

if __name__ == '__main__':
    data = [
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["C", "D", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "CD", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["BD", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ["B", "C", "A", "B", "B", "A", "B", "C", "D", "A", "C", "A", "D", "E", "E", "E", "E", "E", "D", "C"], 
    ]
    run = PyExcel(*data)
    pass

